def test_func(data, template, filename):
  '''
  DOCSTRING: Функція, яка формує авансовий звіт в форматі .xlsx з шаблону template
  INPUT: Масив даних (list), імʼя шаблонного файлу - template (str), імʼя вихідного файлу - filename (str) 
  OUTPUT: збережено готовий файл з заданим імʼям - filename
  '''
  import openpyxl
  from openpyxl.styles import NamedStyle, Font, Border, Side
  from copy import copy
  from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
  from openpyxl import Workbook
  from openpyxl.worksheet.dimensions import ColumnDimension
  from openpyxl.worksheet.dimensions import RowDimension
  from openpyxl.worksheet.page import PrintPageSetup
  from openpyxl.worksheet.pagebreak import Break
  import locale
  import datetime
  from datetime import datetime

  # Завантажуємо шаблонний файл
  wb1 = openpyxl.load_workbook(filename=template)  # Відкриваємо шаблонний файл template
  # Відкриваємо активний лист
  sheet1 = wb1.active

  # Створюємо новий файл
  wb2 = openpyxl.Workbook()
  # Додаємо новий лист
  wb2.create_sheet(title = 'Первый лист', index = 0)
  # Відкриваємо лист
  sheet2 = wb2['Первый лист']

  # Копіюємо об'єднання комірок
  for items in sorted(sheet1.merged_cells.ranges):
    sheet2.merge_cells(str(items))

  # Копіюємо висоту рядків
  for k, cd in sheet1.row_dimensions.items():
    sheet2.row_dimensions[k].height = cd.height 

  # Копіюємо ширину рядків
  for k, cd in sheet1.column_dimensions.items():
    sheet2.column_dimensions[str(k)].width = (cd.width*0.795)
  sheet2.column_dimensions['J'].width = 3.405    

  # Копіюємо форматування комірок
  for row in sheet1.rows:
      for cell in row:
          new_cell = sheet2.cell(row=cell.row, column=cell.column,
                  value= cell.value)
          if cell.has_style:
              new_cell.font = copy(cell.font)
              new_cell.border = copy(cell.border)
              new_cell.fill = copy(cell.fill)
              new_cell.number_format = copy(cell.number_format)
              new_cell.protection = copy(cell.protection)
              new_cell.alignment = copy(cell.alignment)            

  # Створюємо іменований стиль
  ns = NamedStyle(name='left_border')
  ns.font = copy(sheet2['P63'].font)
  ns.alignment = copy(sheet2['P63'].alignment)
  border = Side(style='thin', color='000000')
  ns.border = Border(left=border, top=border, right=border, bottom=border)

  # Додаємо іменований стиль до активного листа 
  wb2.add_named_style(ns) 

  # Додаємо іменований стиль 'left_border' до діапазону комірок 'H24:J39' (чомусь при копіювання не відбулося в даному діапазоні)
  for cell in sheet2['H24:J39']:
    for el in cell:
      el.style = 'left_border'

  # Додаємо іменований стиль 'left_border' до діапазону комірок 'Q63:S74' (чомусь при копіювання не відбулося в даному діапазоні)
  for cell in sheet2['Q63:S74']:
    for el in cell:
      el.style = 'left_border'

  locale.setlocale(locale.LC_ALL, 'uk_UA.UTF-8')
  myDate = datetime.strptime(data[1], "%d.%m.%y")
  data_date_of_report = myDate.strftime('%d %B')
  data_name = data[0]
  data_name_latin = data[8]
  data_ind_number = data[9]
  data_code = data[2]
  data_days = data[5]
  data_pay_day = round(float(data[6]), 2)
  description_of_day = f'Добові ({data_days}х{data_pay_day} грн)'
  sheet2['E64'].value = description_of_day

  payments = data[10]
  row_count = 65
  for payment in payments:
      sheet2['C'+ str(row_count)].value = payment[1]
      sheet2['E'+ str(row_count)].value = payment[2]
      sheet2['P'+ str(row_count)].value = float(payment[3])
      row_count += 1 

  payments_by_card = data[11]
  row_count = 29
  num_row = 1
  for payment in payments_by_card:
      sheet2['A'+ str(row_count)].value = f'{num_row}. Корп. Рах {payment[1]}'
      sheet2['H'+ str(row_count)].value = float(payment[3])
      row_count += 1 
      num_row += 1


  sheet2['B17'].value = data_name
  sheet2['F16'].value = data_name_latin
  sheet2['O6'].value = data_date_of_report
  sheet2['M6'].value = data_code
  sheet2['W57'].value = data_days
  sheet2['V57'].value = data_pay_day

  # Додавання розриву сторінки після 60-го ряду
  row_number = 60  # the row that you want to insert page break
  page_break = Break(id=row_number)  # create Break obj
  sheet2.row_breaks.append(page_break)

  # Встановлюємо діапазон друку
  sheet2.print_area = 'A1:S77'

  # Налаштування друку
  ps = PrintPageSetup()
  # Розмір сторінки
  ps.paperSize = '9' # A4
  sheet2.page_setup = ps
  # Зумувати по ширині
  sheet2.sheet_properties.pageSetUpPr.fitToPage = True
  sheet2.page_setup.fitToHeight = False

  # Зберігання файлу
  wb2.save(filename)
  # Закриваємо шаблонний файл 
  wb1.close()

  # Закриваємо цільовий файл
  wb2.close()